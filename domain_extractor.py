"""Module trích xuất tên miền từ kết quả tìm kiếm"""
import json
import time
import random
import urllib.parse
from urllib.parse import urlparse
import logging
from pathlib import Path
from typing import List, Dict, Any, Set
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
from config import PROFILE_PATH, CHROME_PATH
from constants import BROWSER_CONFIG, TIMEOUTS, DELAYS
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger()


def _extract_domains_from_search_results(page, keyword: str, search_engine: str = "baidu") -> Set[str]:
    """Trích xuất tất cả domains từ kết quả tìm kiếm"""
    domains = set()

    if search_engine == "baidu":
        result_selectors = [
            "div.c-result.result",
            "div.c-result",
            "article.c-container",
        ]
    else:  # Google
        result_selectors = [
            "div.g",
            "div[data-ved]",
            ".tF2Cxc",
        ]

    # Tìm containers
    containers = []
    for selector in result_selectors:
        containers = page.query_selector_all(selector)
        if containers:
            break

    for container in containers:
        domain = ""

        if search_engine == "baidu":
            try:
                data_log = container.get_attribute("data-log")
                if data_log:
                    info = json.loads(data_log)
                    mu = info.get("mu") or ""
                    if mu:
                        domain = urlparse(mu).netloc
            except Exception:
                pass
        else:  # Google
            try:
                cite_elem = container.query_selector("cite")
                if cite_elem:
                    cite_text = cite_elem.text_content().strip()
                    if cite_text:
                        if not cite_text.startswith(("http://", "https://")):
                            cite_text = "https://" + cite_text
                        domain = urlparse(cite_text).netloc
                else:
                    link_elem = container.query_selector("a[href]")
                    if link_elem:
                        href = link_elem.get_attribute("href")
                        if href and href.startswith("http"):
                            domain = urlparse(href).netloc
            except Exception:
                pass

        if domain and domain not in {"www.baidu.com", "www.google.com"}:
            domains.add(domain)

    return domains


def _navigate_with_retry(page, url: str, keyword: str) -> bool:
    """Điều hướng đến URL với retry logic"""
    for retry in range(2):
        try:
            wait_until = "domcontentloaded" if retry == 0 else "load"
            timeout = TIMEOUTS["navigation"] if retry == 0 else TIMEOUTS["navigation_fallback"]
            page.goto(url, wait_until=wait_until, timeout=timeout)
            return True
        except PlaywrightTimeoutError:
            if retry == 1:
                try:
                    page.goto(url, wait_until="commit", timeout=TIMEOUTS["navigation_commit"])
                    page.wait_for_timeout(2000)
                    logger.warning(f"[⚠️] Dùng fallback navigation cho: {keyword}")
                    return True
                except Exception:
                    pass
            else:
                time.sleep(0.5)
    return False


def _search_and_extract_domains(
    keywords: List[str], search_engine: str = "baidu", page=None
) -> Dict[str, Set[str]]:
    """Trích xuất domains cho danh sách từ khóa, tái sử dụng page nếu được truyền vào"""
    results = {}

    for idx, keyword in enumerate(keywords, 1):
        kw = str(keyword).strip()
        if not kw:
            continue

        logger.info(f"[🔍] ({idx}/{len(keywords)}) Trích xuất domains từ {search_engine.upper()}: {kw}")

        try:
            encoded_kw = urllib.parse.quote(kw, encoding="utf-8")
            if search_engine == "baidu":
                url = f"https://m.baidu.com/s?word={encoded_kw}"
            else:
                url = f"https://www.google.com/search?q={encoded_kw}"

            if not _navigate_with_retry(page, url, kw):
                logger.error(f"   ❌ Không thể điều hướng đến trang tìm kiếm cho: {kw}")
                results[kw] = set()
                continue

            # Chờ kết quả xuất hiện thay vì sleep cứng
            result_selectors = (
                "div.c-result.result, div.c-result, article.c-container"
                if search_engine == "baidu"
                else "div.g, div[data-ved], .tF2Cxc"
            )
            try:
                page.wait_for_selector(result_selectors, timeout=TIMEOUTS["selector_visible"], state="visible")
            except PlaywrightTimeoutError:
                # Fallback: chờ attached
                try:
                    page.wait_for_selector(result_selectors, timeout=TIMEOUTS["selector_attached"], state="attached")
                except PlaywrightTimeoutError:
                    logger.warning(f"   ⚠️ Không tìm thấy kết quả cho: {kw}")

            domains = _extract_domains_from_search_results(page, kw, search_engine)
            results[kw] = domains

            logger.info(
                f"   ✅ Tìm thấy {len(domains)} domains: "
                f"{', '.join(list(domains)[:5])}{'...' if len(domains) > 5 else ''}"
            )

        except Exception as e:
            logger.error(f"   ❌ Lỗi khi xử lý {kw}: {e}")
            results[kw] = set()

        delay = random.uniform(DELAYS["normal_min"], DELAYS["normal_max"])
        time.sleep(delay)

    return results


def _format_excel(ws, num_cols: int) -> None:
    """Áp dụng format chung cho worksheet: header, auto-width, autofilter"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for col_idx in range(1, num_cols + 1):
        max_width = 0
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                max_width = max(max_width, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(
            max(max_width + 2, 15), 80
        )

    if ws.max_row > 1:
        last_col_letter = ws.cell(row=1, column=num_cols).column_letter
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"


def _save_excel_with_format(df: pd.DataFrame, output_file: str) -> None:
    """Ghi DataFrame vào Excel và áp dụng format"""
    df.to_excel(output_file, index=False)
    try:
        wb = load_workbook(output_file)
        ws = wb.active
        _format_excel(ws, len(df.columns))
        wb.save(output_file)
    except Exception as e:
        logger.warning(f"[⚠️] Không thể format Excel: {e}")


def extract_domains_to_excel(keywords: List[str], output_file: str = "domains_extracted.xlsx") -> None:
    """Trích xuất domains từ cả Baidu và Google, ghi vào Excel.
    Dùng một browser session duy nhất cho cả hai engine."""
    if not keywords:
        logger.error("❌ Không có từ khóa để xử lý!")
        return

    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            user_data_dir=PROFILE_PATH,
            headless=False,
            executable_path=CHROME_PATH,
            **BROWSER_CONFIG,
        )
        page = browser.new_page()

        logger.info("🔍 Bắt đầu trích xuất domains từ Baidu...")
        baidu_results = _search_and_extract_domains(keywords, "baidu", page)

        logger.info("🔍 Bắt đầu trích xuất domains từ Google...")
        google_results = _search_and_extract_domains(keywords, "google", page)

        browser.close()

    # Tạo DataFrame kết quả
    data = []
    for keyword in keywords:
        kw = str(keyword).strip()
        if not kw:
            continue

        baidu_domains = list(baidu_results.get(kw, set()))
        google_domains = list(google_results.get(kw, set()))
        all_domains = list(set(baidu_domains + google_domains))

        data.append({
            "Từ khóa": kw,
            "Domains từ Baidu": ", ".join(baidu_domains),
            "Domains từ Google": ", ".join(google_domains),
            "Tất cả Domains": ", ".join(all_domains),
            "Số lượng Baidu": len(baidu_domains),
            "Số lượng Google": len(google_domains),
            "Tổng số domains": len(all_domains),
        })

    _save_excel_with_format(pd.DataFrame(data), output_file)

    # Thống kê
    total_keywords = len([kw for kw in keywords if str(kw).strip()])
    total_baidu = sum(len(d) for d in baidu_results.values())
    total_google = sum(len(d) for d in google_results.values())

    logger.info("\n" + "=" * 60)
    logger.info("📊 THỐNG KÊ TRÍCH XUẤT DOMAINS")
    logger.info("=" * 60)
    logger.info(f"📝 Tổng số từ khóa: {total_keywords}")
    logger.info(f"🔍 Tổng domains từ Baidu: {total_baidu}")
    logger.info(f"🔍 Tổng domains từ Google: {total_google}")
    logger.info(f"📄 Đã lưu kết quả vào: {output_file}")
    logger.info("✅ Hoàn thành trích xuất domains.")
    logger.info("=" * 60)


def extract_top_domains(
    keywords: List[str], top_n: int = 10, output_file: str = "top_domains.xlsx"
) -> None:
    """Trích xuất top domains xuất hiện nhiều nhất.
    Dùng một browser session duy nhất cho cả hai engine."""
    if not keywords:
        logger.error("❌ Không có từ khóa để xử lý!")
        return

    logger.info("🔍 Bắt đầu trích xuất top domains...")

    with sync_playwright() as p:
        browser = p.chromium.launch_persistent_context(
            user_data_dir=PROFILE_PATH,
            headless=False,
            executable_path=CHROME_PATH,
            **BROWSER_CONFIG,
        )
        page = browser.new_page()

        baidu_results = _search_and_extract_domains(keywords, "baidu", page)
        google_results = _search_and_extract_domains(keywords, "google", page)

        browser.close()

    # Đếm tần suất xuất hiện
    domain_count: Dict[str, int] = {}
    domain_keywords: Dict[str, List[str]] = {}

    for keyword in keywords:
        kw = str(keyword).strip()
        if not kw:
            continue

        all_domains = set()
        all_domains.update(baidu_results.get(kw, set()))
        all_domains.update(google_results.get(kw, set()))

        for domain in all_domains:
            domain_count[domain] = domain_count.get(domain, 0) + 1
            domain_keywords.setdefault(domain, []).append(kw)

    sorted_domains = sorted(domain_count.items(), key=lambda x: x[1], reverse=True)

    data = []
    for domain, count in sorted_domains[:top_n]:
        kw_list = domain_keywords[domain]
        data.append({
            "Domain": domain,
            "Số lần xuất hiện": count,
            "Từ khóa tìm thấy": ", ".join(kw_list[:5]) + ("..." if len(kw_list) > 5 else ""),
            "Tổng số từ khóa": len(kw_list),
        })

    _save_excel_with_format(pd.DataFrame(data), output_file)

    logger.info("\n" + "=" * 60)
    logger.info(f"📊 TOP {top_n} DOMAINS PHỔ BIẾN NHẤT")
    logger.info("=" * 60)
    for i, (domain, count) in enumerate(sorted_domains[:top_n], 1):
        logger.info(f"{i:2d}. {domain} ({count} lần)")
    logger.info(f"📄 Đã lưu kết quả vào: {output_file}")
    logger.info("=" * 60)

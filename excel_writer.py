"""Module ghi kết quả tìm kiếm vào Excel"""
import pandas as pd
import logging
from pathlib import Path
from typing import List, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText, InlineFont
from config import EXCEL_FILE

logger = logging.getLogger()


def _calculate_text_width(text: str) -> float:
    """Tính độ rộng của text, xử lý ký tự Unicode"""
    width = 0
    for char in text:
        width += 1.0 if ord(char) < 128 else 2.0
    return width


def _read_existing_main_titles(excel_path: Path) -> Tuple[dict, list]:
    """Đọc cột 'Tiêu đề chính' từ file Excel hiện có."""
    by_keyword, by_row = {}, []
    if not excel_path.exists():
        return by_keyword, by_row

    try:
        wb = load_workbook(str(excel_path), read_only=True, data_only=True)
        ws = wb.active
        main_title_col = None
        
        for col_idx in range(1, (ws.max_column or 0) + 1):
            if str(ws.cell(row=1, column=col_idx).value or "").strip() == "Tiêu đề chính":
                main_title_col = col_idx
                break

        if main_title_col:
            for row_idx in range(2, (ws.max_row or 1) + 1):
                keyword_cell = ws.cell(row=row_idx, column=1).value
                main_title_cell = ws.cell(row=row_idx, column=main_title_col).value
                val = str(main_title_cell).strip() if main_title_cell else ""
                by_row.append(val)
                if keyword_cell and val:
                    by_keyword[str(keyword_cell).strip()] = val

        wb.close()
    except Exception as e:
        logger.warning(f"[⚠️] Không thể đọc cột 'Tiêu đề chính': {e}")

    return by_keyword, by_row


def _write_excel_full(
    out_df: pd.DataFrame,
    path: Path,
    added_texts: List[str],
    existing_main_titles: dict,
) -> Path:
    """Ghi DataFrame vào Excel với formatting."""
    tmp_path = path.with_suffix(".tmp.xlsx")
    out_df.to_excel(str(tmp_path), index=False, header=False)

    try:
        wb = load_workbook(str(tmp_path))
        ws = wb.active

        # Khôi phục cột "Tiêu đề chính"
        if existing_main_titles:
            for row_idx in range(2, ws.max_row + 1):
                keyword_cell = ws.cell(row=row_idx, column=1).value
                if keyword_cell:
                    kw_str = str(keyword_cell).strip()
                    if kw_str in existing_main_titles:
                        ws.cell(row=row_idx, column=6).value = existing_main_titles[kw_str]

        # Format header
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for col_idx in range(1, 7):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # AutoFilter
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:F{ws.max_row}"

        # Căn giữa cột STT
        center_alignment = Alignment(horizontal="center", vertical="center")
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=5).alignment = center_alignment

        # Tô đỏ keyword và added_text
        red_font_obj = InlineFont(rFont="Calibri", color="FFFF0000")
        for row_idx in range(2, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1)
            cell_b = ws.cell(row=row_idx, column=2)

            title_original = str(cell_b.value).strip() if cell_b.value else ""
            keyword = str(cell_a.value).strip() if cell_a.value else ""
            title_starts_with_keyword = bool(keyword and title_original.startswith(keyword))

            # Tô đỏ keyword trong cột A
            if title_starts_with_keyword:
                current_font = cell_a.font or Font()
                cell_a.font = Font(
                    name=current_font.name, size=current_font.size,
                    bold=current_font.bold, italic=current_font.italic,
                    underline=current_font.underline, strike=current_font.strike,
                    color="FF0000",
                )

            # Tô đỏ trong title
            if cell_b.value:
                added_text = added_texts[row_idx - 2] if (row_idx - 2) < len(added_texts) else ""
                try:
                    rich_text = CellRichText()
                    parts_to_highlight = []

                    if title_starts_with_keyword and keyword:
                        parts_to_highlight.append((0, len(keyword), keyword))

                    if added_text and added_text in title_original:
                        idx = title_original.rfind(added_text)
                        if idx >= 0:
                            parts_to_highlight.append((idx, idx + len(added_text), added_text))

                    if parts_to_highlight:
                        parts_to_highlight.sort(key=lambda x: x[0])
                        current_pos = 0
                        for start, end, text in parts_to_highlight:
                            if start > current_pos:
                                before = title_original[current_pos:start]
                                if before:
                                    rich_text.append(before)
                            rich_text.append(TextBlock(red_font_obj, text))
                            current_pos = end
                        if current_pos < len(title_original):
                            after = title_original[current_pos:]
                            if after:
                                rich_text.append(after)
                        if len(rich_text) > 0:
                            cell_b.value = rich_text
                    elif title_starts_with_keyword:
                        rich_text.append(TextBlock(red_font_obj, keyword))
                        remaining = title_original[len(keyword):]
                        if remaining:
                            rich_text.append(remaining)
                        cell_b.value = rich_text
                except Exception as e:
                    logger.debug(f"Lỗi tô màu hàng {row_idx}: {e}")

        # Tự động điều chỉnh độ rộng cột
        for col_idx in range(1, 7):
            letter = get_column_letter(col_idx)
            max_width = 0
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.value:
                    cell_text = "".join(str(part) for part in cell.value) if hasattr(cell.value, "__iter__") and not isinstance(cell.value, str) else str(cell.value)
                    text_width = _calculate_text_width(cell_text) + 3
                    max_width = max(max_width, text_width)
            width = min(max(max_width, 10), 100) if max_width > 0 else 15
            ws.column_dimensions[letter].width = width

        wb.save(str(path))
        wb.close()
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

    return path


def _safe_write_excel(
    out_df: pd.DataFrame,
    path: str,
    added_texts: List[str],
    make_backup: bool = False,
    existing_main_titles: dict = None,
) -> str:
    """Ghi Excel an toàn với fallback nếu file đang mở"""
    if existing_main_titles is None:
        existing_main_titles = {}

    primary_path = Path(path)
    try:
        written = _write_excel_full(out_df, primary_path, added_texts, existing_main_titles)
    except PermissionError:
        backup_path = primary_path.with_name(primary_path.stem + "_output.xlsx")
        logger.warning(f"[⚠️] File '{primary_path}' đang mở. Lưu vào: {backup_path}")
        written = _write_excel_full(out_df, backup_path, added_texts, existing_main_titles)

    if make_backup:
        backup_path = primary_path.with_name(primary_path.stem + "_output.xlsx")
        if written.resolve() != backup_path.resolve():
            try:
                _write_excel_full(out_df, backup_path, added_texts, existing_main_titles)
            except PermissionError:
                pass

    return str(written)


def write_search_results(
    keywords: List[str],
    results: List[Tuple[str, str, str, bool, str, str]],
    duplicate_count: int,
    success_count: int,
    error_count: int,
    total: int,
) -> None:
    """Ghi kết quả tìm kiếm vào Excel"""
    excel_path = Path(EXCEL_FILE)
    existing_main_titles_by_kw, existing_main_titles_by_row = _read_existing_main_titles(excel_path)

    # Tạo DataFrame
    output_df = pd.DataFrame(index=range(len(results)))
    for col in range(4):
        output_df[col] = ""

    added_texts = []
    for i, (title, domain, time_tag, is_processed, added_text, original_title) in enumerate(results):
        keyword_str = keywords[i].strip() if i < len(keywords) and keywords[i] else ""
        output_df.iloc[i, 0] = keyword_str
        output_df.iloc[i, 1] = str(title).strip() if title else ""
        output_df.iloc[i, 2] = str(domain).strip() if domain else ""
        output_df.iloc[i, 3] = str(time_tag).strip() if time_tag else ""
        added_texts.append(added_text)

    output_df[4] = [i + 1 for i in range(len(results))]

    # Cột "Tiêu đề chính"
    main_titles = list(existing_main_titles_by_row)
    length_gap = len(results) - len(main_titles)
    if length_gap > 0:
        main_titles += [""] * length_gap
    elif length_gap < 0:
        main_titles = main_titles[: len(results)]
    output_df[5] = main_titles

    # Thêm header
    header_row = pd.DataFrame([["Từ khóa", "Tiêu đề", "Domain", "Thời gian", "STT", "Tiêu đề chính"]])
    output_df = pd.concat([header_row, output_df], ignore_index=True)

    try:
        written_path = _safe_write_excel(
            output_df, EXCEL_FILE, added_texts,
            make_backup=error_count > 0,
            existing_main_titles=existing_main_titles_by_kw,
        )
        if written_path != EXCEL_FILE:
            logger.warning(f"[⚠️] Lưu vào: {written_path}")
        logger.info("✅ Hoàn tất ghi file.")
    except Exception as e:
        logger.error(f"[❌] Lỗi khi ghi file: {e}", exc_info=True)
    finally:
        logger.info("\n" + "=" * 60)
        logger.info("📊 THỐNG KÊ KẾT QUẢ")
        logger.info("=" * 60)
        logger.info(f"📝 Tổng số từ khóa: {total}")
        logger.info(f"🔄 Số trùng lặp: {duplicate_count}")
        logger.info(f"✅ Số thành công: {success_count}")
        logger.info(f"❌ Số lỗi: {error_count}")
        logger.info("=" * 60)

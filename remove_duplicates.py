"""Module loại bỏ từ khóa trùng lặp"""
import pandas as pd
import logging
from pathlib import Path
from typing import List, Tuple, Dict
from collections import defaultdict
from openpyxl import load_workbook
from config import EXCEL_FILE

logger = logging.getLogger()


def remove_duplicates_and_color() -> None:
    """Loại bỏ từ khóa trùng lặp (giữ lại lần đầu tiên)"""
    if not Path(EXCEL_FILE).exists():
        logger.error(f"❌ File {EXCEL_FILE} không tồn tại!")
        return

    # Đọc file Excel
    try:
        df = pd.read_excel(EXCEL_FILE, header=None, dtype=str)
    except Exception as e:
        logger.error(f"❌ Lỗi khi đọc file Excel: {e}", exc_info=True)
        return

    if df.empty or len(df.columns) < 1:
        logger.error("❌ File Excel trống hoặc không có cột A!")
        return

    # Đọc từ khóa từ cột A (bắt đầu từ row 2, index 1)
    # Loại bỏ khoảng trống ở đầu/cuối và thay tab bằng space
    keywords: List[Tuple[int, str]] = []
    for idx in range(1, len(df)):
        kw_raw = str(df.iloc[idx, 0]) if pd.notna(df.iloc[idx, 0]) else ""
        kw = kw_raw.strip()
        if kw:
            keywords.append((idx + 1, kw))  # Lưu (row_number, keyword)

    if not keywords:
        logger.error("❌ Không có từ khóa hợp lệ!")
        return

    # Tìm các từ khóa trùng lặp
    keyword_groups: Dict[str, List[int]] = defaultdict(list)
    for row_num, kw in keywords:
        keyword_groups[kw].append(row_num)

    duplicates = {kw: rows for kw, rows in keyword_groups.items() if len(rows) > 1}

    if duplicates:
        logger.info(f"📊 Tìm thấy {len(duplicates)} từ khóa trùng lặp:")
        for kw, rows in duplicates.items():
            logger.info(
                f"  - '{kw}': {len(rows)} lần (hàng {', '.join(map(str, rows))})"
            )
    else:
        logger.info("✅ Không có từ khóa trùng lặp!")

    # Loại bỏ các từ khóa trùng lặp (giữ lại lần đầu tiên) và làm sạch dữ liệu
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        rows_to_remove = []
        if duplicates:
            seen_keywords = set()
            for row_num, kw in keywords:
                if kw in seen_keywords:
                    rows_to_remove.append(row_num)
                else:
                    seen_keywords.add(kw)

            # Xóa các hàng trùng lặp (từ dưới lên để không ảnh hưởng index)
            for row_num in sorted(rows_to_remove, reverse=True):
                ws.delete_rows(row_num)

        # Làm sạch dữ liệu: loại bỏ space/tab ở đầu/cuối cho tất cả các hàng còn lại
        # Bắt đầu từ row 2 (index 1) vì row 1 là header
        cleaned_count = 0
        for row_idx in range(2, ws.max_row + 1):
            # Làm sạch cột A (từ khóa)
            cell_a = ws.cell(row=row_idx, column=1)
            if cell_a.value:
                kw_original = str(cell_a.value)
                kw_cleaned = kw_original.strip()
                if kw_original != kw_cleaned:
                    cell_a.value = kw_cleaned
                    cleaned_count += 1

            # Làm sạch cột B (title) nếu có dữ liệu
            cell_b = ws.cell(row=row_idx, column=2)
            if cell_b.value:
                title_original = str(cell_b.value)
                title_cleaned = title_original.strip()
                if title_original != title_cleaned:
                    cell_b.value = title_cleaned
                    cleaned_count += 1

        wb.save(EXCEL_FILE)
        logger.info(f"✅ Đã loại bỏ {len(rows_to_remove)} từ khóa trùng lặp!")
        if cleaned_count > 0:
            logger.info(
                f"🧹 Đã làm sạch {cleaned_count} ô dữ liệu (loại bỏ space/tab ở đầu/cuối)!"
            )
        logger.info(f"📝 Đã lưu vào file: {EXCEL_FILE}")

    except Exception as e:
        logger.error(f"❌ Lỗi khi xử lý file Excel: {e}", exc_info=True)

"""Module lọc từ khóa theo đài"""
import pandas as pd
import logging
from pathlib import Path
from typing import List, Dict, Any
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config import EXCEL_FILE

logger = logging.getLogger()

# Danh sách các đài cần kiểm tra (không phân biệt hoa thường)
# Format: {"Tên đài hiển thị": ["từ khóa 1", "từ khóa 2", ...]}
PLATFORMS = {
    "爱游戏": ["AYX", "爱游戏"],
    "乐鱼": ["LEYU", "乐鱼"],
    "华体会": ["HTH", "华体会"],
    "米兰": ["MILAN", "米兰"],
    "星空": ["XINGKONG", "星空"],
    "乐竞": ["LEJING", "乐竞"],
    "九游": ["JY", "九游"],
    "开云": ["KY", "开云"],
    "MK": ["MK"],
}


def _read_excel_rows() -> List[Dict[str, Any]]:
    """Đọc tất cả các dòng từ file Excel (bỏ qua header)"""
    if not Path(EXCEL_FILE).exists():
        logger.error(f"❌ File {EXCEL_FILE} không tồn tại!")
        return []

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except Exception as e:
        logger.error(f"❌ Không thể đọc file Excel: {e}")
        return []

    all_rows = []
    for row_idx in range(2, ws.max_row + 1):
        keyword_raw = ws.cell(row=row_idx, column=1).value
        keyword = str(keyword_raw).strip() if keyword_raw else ""

        if keyword:  # Chỉ xử lý nếu có từ khóa
            row_data = {"row_idx": row_idx, "keyword": keyword}
            # Lấy tất cả các cột
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                row_data[f"col_{col_idx}"] = (
                    str(cell_value).strip() if cell_value else ""
                )
            # Lấy title từ cột B (nếu có)
            title_raw = (
                ws.cell(row=row_idx, column=2).value
                if ws.max_column >= 2
                else None
            )
            row_data["title"] = str(title_raw).strip() if title_raw else ""
            all_rows.append(row_data)

    return all_rows


def _check_platform_match(keyword: str, title: str, check_title: bool) -> bool:
    """Kiểm tra từ khóa hoặc title có khớp với từ khóa đài không"""
    keyword_lower = keyword.lower()
    title_lower = title.lower() if title else ""

    for platform_name, keywords_list in PLATFORMS.items():
        for platform_keyword in keywords_list:
            platform_lower = platform_keyword.lower()
            # Kiểm tra trong cột A (keyword) hoặc cột B (title) nếu check_title=True
            if platform_lower in keyword_lower:
                return True
            if check_title and platform_lower in title_lower:
                return True

    return False


def _remove_duplicates(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Loại bỏ trùng lặp dựa trên cột A (giữ lại lần đầu tiên)"""
    seen_keywords = set()
    unique_rows = []
    duplicate_count = 0

    for row in rows:
        keyword = row["keyword"]
        if keyword not in seen_keywords:
            seen_keywords.add(keyword)
            unique_rows.append(row)
        else:
            duplicate_count += 1

    if duplicate_count > 0:
        logger.info(
            f"🔄 Đã loại bỏ {duplicate_count} từ khóa trùng lặp (dựa trên cột A)"
        )

    return unique_rows


def _write_results_to_excel(
    unique_rows: List[Dict[str, Any]], ws_original, num_cols: int
) -> None:
    """Ghi kết quả vào file Excel với format"""
    output_path = Path(EXCEL_FILE)

    try:
        # Tạo DataFrame mới
        result_data = {}

        # Đọc header từ file gốc
        headers = []
        for col_idx in range(1, num_cols + 1):
            header_value = ws_original.cell(row=1, column=col_idx).value
            headers.append(str(header_value) if header_value else f"Cột {col_idx}")

        # Tạo dictionary với các cột
        for col_idx in range(1, num_cols + 1):
            col_name = (
                headers[col_idx - 1]
                if col_idx <= len(headers)
                else f"Cột {col_idx}"
            )
            result_data[col_name] = [
                row.get(f"col_{col_idx}", "") for row in unique_rows
            ]

        df_result = pd.DataFrame(result_data)

        # Ghi đè lên file Excel gốc
        df_result.to_excel(output_path, index=False, engine="openpyxl")

        # Format lại file
        wb_result = load_workbook(output_path)
        ws_result = wb_result.active

        # Header style
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Format header
        for col_idx in range(1, num_cols + 1):
            cell = ws_result.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Hàm tính độ rộng text, xử lý ký tự Unicode
        def _calculate_text_width(text: str) -> float:
            width = 0
            for char in text:
                # Ký tự ASCII thường: ~1 đơn vị
                # Ký tự Unicode (tiếng Trung, Nhật, Hàn): ~2 đơn vị
                if ord(char) < 128:
                    width += 1.0
                else:
                    width += 2.0
            return width
        
        # Tự động điều chỉnh độ rộng cột theo nội dung
        for col_idx in range(1, num_cols + 1):
            letter = get_column_letter(col_idx)
            max_width = 0
            
            # Duyệt qua tất cả các cell trong cột (bao gồm header)
            for row in ws_result.iter_rows(min_row=1, max_row=ws_result.max_row, min_col=col_idx, max_col=col_idx):
                cell = row[0]
                if cell.value:
                    # Tính độ dài của text
                    if hasattr(cell.value, '__iter__') and not isinstance(cell.value, str):
                        # Rich text - tính tổng độ dài của tất cả các phần
                        cell_text = ''.join(str(part) for part in cell.value)
                    else:
                        cell_text = str(cell.value)
                    
                    # Tính độ rộng thực tế (xử lý Unicode)
                    text_width = _calculate_text_width(cell_text)
                    # Cộng thêm padding (2-3 ký tự)
                    text_width = text_width + 3
                    max_width = max(max_width, text_width)
            
            # Đặt độ rộng tối thiểu và tối đa
            if max_width > 0:
                width = min(max(max_width, 10), 100)  # Tối thiểu 10, tối đa 100
            else:
                width = 15  # Độ rộng mặc định nếu cột trống
            
            ws_result.column_dimensions[letter].width = width

        # Thêm AutoFilter
        if ws_result.max_row > 1:
            ws_result.auto_filter.ref = (
                f"A1:{get_column_letter(num_cols)}{ws_result.max_row}"
            )

        wb_result.save(output_path)
        logger.info(f"✅ Đã cập nhật file: {output_path}")

    except Exception as e:
        logger.error(f"❌ Lỗi khi ghi file Excel: {e}", exc_info=True)
        raise


def _filter_keywords_common(check_title: bool) -> None:
    """Hàm chung để lọc từ khóa theo đài"""
    if not Path(EXCEL_FILE).exists():
        logger.error(f"❌ File {EXCEL_FILE} không tồn tại!")
        return

    # Đọc toàn bộ file Excel
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    except Exception as e:
        logger.error(f"❌ Không thể đọc file Excel: {e}")
        return

    # Đọc tất cả các dòng
    all_rows = _read_excel_rows()

    if not all_rows:
        logger.error("❌ Không có từ khóa hợp lệ!")
        return

    logger.info(f"📝 Tổng số dòng ban đầu: {len(all_rows)}")

    # Lọc các dòng khớp với từ khóa đài
    matched_rows = []
    removed_count = 0

    for row in all_rows:
        keyword = row["keyword"]
        title = row.get("title", "")

        if _check_platform_match(keyword, title, check_title):
            matched_rows.append(row)
        else:
            removed_count += 1

    title_info = " (trong cả cột A và B)" if check_title else ""
    logger.info(f"🗑️  Đã xóa {removed_count} dòng không khớp đài nào{title_info}")
    logger.info(f"✅ Giữ lại {len(matched_rows)} dòng khớp")

    # Loại bỏ trùng lặp
    unique_rows = _remove_duplicates(matched_rows)

    logger.info(f"📝 Số dòng cuối cùng: {len(unique_rows)}")

    # Sắp xếp theo từ khóa
    unique_rows.sort(key=lambda x: x["keyword"])

    # Ghi kết quả vào Excel
    try:
        _write_results_to_excel(unique_rows, ws, ws.max_column)

        logger.info("\n" + "=" * 60)
        logger.info("📊 THỐNG KÊ")
        logger.info("=" * 60)
        logger.info(f"📝 Tổng số dòng ban đầu: {len(all_rows)}")
        logger.info(f"🗑️  Đã xóa dòng không khớp: {removed_count}")
        logger.info(
            f"🔄 Đã xóa dòng trùng lặp: {len(matched_rows) - len(unique_rows)}"
        )
        logger.info(f"✅ Số dòng cuối cùng: {len(unique_rows)}")
        logger.info("=" * 60)

    except Exception as e:
        logger.error(f"❌ Lỗi khi xử lý: {e}", exc_info=True)


def filter_keywords_by_platform() -> None:
    """Lọc từ khóa: giữ lại các dòng có từ khóa đài trong cột A, xóa các dòng không khớp, sau đó loại bỏ trùng lặp"""
    _filter_keywords_common(check_title=False)


def filter_keywords_by_platform_with_title() -> None:
    """Lọc từ khóa: giữ lại các dòng có từ khóa đài trong cột A HOẶC cột B, xóa các dòng không khớp, sau đó loại bỏ trùng lặp"""
    _filter_keywords_common(check_title=True)

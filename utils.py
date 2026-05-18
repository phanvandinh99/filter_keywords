"""Module tiện ích để đọc/ghi từ khóa từ Excel"""
import pandas as pd
import logging
from pathlib import Path
from typing import List
from openpyxl import load_workbook, Workbook
from config import EXCEL_FILE

logger = logging.getLogger()


def read_keywords_from_excel() -> List[str]:
    """Đọc từ khóa từ file Excel (hàng 1 là header, đọc từ A2)"""
    if not Path(EXCEL_FILE).exists():
        logger.error(f"❌ File {EXCEL_FILE} không tồn tại!")
        return []

    try:
        df = pd.read_excel(EXCEL_FILE, header=None, dtype=str, skiprows=1)
        if df.empty or len(df.columns) < 1:
            logger.error("❌ File Excel trống hoặc không có cột A!")
            return []

        keywords = []
        for kw in df.iloc[:, 0].fillna("").astype(str):
            kw_clean = str(kw).strip()
            if kw_clean:
                keywords.append(kw_clean)

        if not keywords:
            logger.error("❌ Không có từ khóa hợp lệ để xử lý.")
            return []

        return keywords
    except Exception as e:
        logger.error(f"❌ Lỗi khi đọc file Excel: {e}", exc_info=True)
        return []


def write_keywords_to_excel(keywords: List[str]) -> bool:
    """
    Ghi danh sách từ khóa vào cột A của keywords.xlsx.
    - Hàng 1: header (giữ nguyên nếu đã có, tạo mới nếu chưa có)
    - Từ hàng 2 trở đi: từ khóa
    - Các cột khác (B, C, D...) được xóa sạch để tránh dữ liệu cũ lẫn lộn
    Trả về True nếu thành công.
    """
    if not keywords:
        logger.error("❌ Danh sách từ khóa rỗng, không ghi vào Excel.")
        return False

    excel_path = Path(EXCEL_FILE)

    try:
        if excel_path.exists():
            wb = load_workbook(str(excel_path))
            ws = wb.active
            # Xóa toàn bộ dữ liệu từ hàng 2 trở đi
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            # Đảm bảo header hàng 1 cột A tồn tại
            if not ws.cell(row=1, column=1).value:
                ws.cell(row=1, column=1).value = "Từ khóa"
        else:
            wb = Workbook()
            ws = wb.active
            ws.cell(row=1, column=1).value = "Từ khóa"

        # Ghi từ khóa vào cột A từ hàng 2
        for i, kw in enumerate(keywords, start=2):
            ws.cell(row=i, column=1).value = kw.strip()

        wb.save(str(excel_path))
        logger.info(f"✅ Đã ghi {len(keywords)} từ khóa vào cột A của {EXCEL_FILE}")
        return True

    except PermissionError:
        logger.error(
            f"❌ Không thể ghi vào {EXCEL_FILE} — file đang được mở. "
            "Hãy đóng file Excel rồi thử lại."
        )
        return False
    except Exception as e:
        logger.error(f"❌ Lỗi khi ghi từ khóa vào Excel: {e}", exc_info=True)
        return False
